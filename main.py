import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill


def round_to_nearest(x, base=5):
    return base * round(x/base)

def create_entry(date, description, amount, type, account):
    return {"Date": date, "Description": description, "Amount": amount, "Type": type, "Account": account}

entries = []

# Define your starting balances
balances = {
    "Checking": 5542.10,
    "Savings": 0,
    "Investing": 0,
    "Taxes": 355.77
}

# Define your bills and paycheck information
bills = [
    {"Description": "Rent", "Amount": -1450, "Due Day": 1}
]
initial_paycheck_date = datetime(2024, 3, 15)
months_to_project = 3
start_date = datetime.today()
end_date = start_date + timedelta(days=months_to_project * 30)

# Process bills
for bill in bills:
    current_date = start_date.replace(day=bill["Due Day"])
    while current_date <= end_date:
        if current_date >= start_date:
            entries.append(create_entry(current_date, bill["Description"], bill["Amount"], "Expense", "Checking"))
        current_date += timedelta(days=30)

# Process paychecks
current_date = initial_paycheck_date
while current_date <= end_date:
    if current_date >= start_date:
        driveline_paycheck = 924.89
        academy_paycheck = 294  # Placeholder for variable paycheck amount
        set_for_taxes = round_to_nearest(academy_paycheck * 0.2)

        # Add the driveline paycheck entry
        entries.append(create_entry(current_date, "Driveline Paycheck", driveline_paycheck, "Income", "Checking"))
        # Add the academy paycheck and tax deduction entries
        entries.append(create_entry(current_date, "Academy Paycheck", academy_paycheck - set_for_taxes, "Income", "Checking"))
        entries.append(create_entry(current_date, "Tax Deductions", set_for_taxes, "Transfer", "Taxes"))


    current_date += timedelta(days=14)


df = pd.DataFrame(entries)
df = df.sort_values(by="Date")

def last_day_of_month(any_day):
    next_month = any_day.replace(day=28) + timedelta(days=4)
    return next_month - timedelta(days=next_month.day)

for account in ["Savings", "Investing", "Taxes"]:
    df[f'{account} Balance'] = None
    last_day_of_prev_month = None

    for i, row in df.iterrows():
        last_day = last_day_of_month(row['Date'])

        if last_day != last_day_of_prev_month:
            df.at[i, f'{account} Balance'] = balances[account]
            last_day_of_prev_month = last_day

# Fill forward the last known balance for each account
for account in ["Savings", "Investing", "Taxes"]:
    df[f'{account} Balance'].ffill(inplace=True)

# Checking balance will update with every transaction
running_checking_balance = balances['Checking']
for i, row in df.iterrows():
    if row['Account'] == 'Checking':
        running_checking_balance += row['Amount']
    df.at[i, 'Checking Balance'] = running_checking_balance

# Save the DataFrame to an Excel file
excel_file_path = "financial_projection.xlsx"
df.to_excel(excel_file_path, index=False)

wb = openpyxl.load_workbook(excel_file_path)
ws = wb.active

green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
blue_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')

for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        # Highlight amounts based on positive or negative values
        if cell.column == 3:  # Assuming the 'Amount' is in the third column
            if cell.value < 0:
                cell.fill = red_fill
            elif cell.value > 0:
                cell.fill = green_fill

        # Highlight types based on Income, Expense, or Transfer
        if cell.column == 4:  # Assuming the 'Type' is in the fourth column
            if cell.value == 'Income':
                cell.fill = green_fill
            elif cell.value == 'Expense':
                cell.fill = red_fill
            elif cell.value == 'Transfer':
                cell.fill = blue_fill

# Save the workbook
wb.save(excel_file_path)